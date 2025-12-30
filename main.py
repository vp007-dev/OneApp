# main.py
import os
import io
import json
import re
import time
import requests
from datetime import date
import datetime
import firebase_admin
from firebase_admin import credentials, messaging
from typing import Dict, Any

from fastapi import (
    FastAPI,
    Request,
    Depends,
    Form,
    UploadFile,
    File,
    HTTPException,
    status,
)
from fastapi.responses import (
    HTMLResponse,
    RedirectResponse,
    JSONResponse,
    FileResponse,
)
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates

from sqlmodel import SQLModel, Session, create_engine, select
from apscheduler.schedulers.background import BackgroundScheduler
from pathlib import Path

import pytesseract
from PIL import Image
# import pdf2image
import openpyxl
from openpyxl.utils.datetime import from_excel as openpyxl_from_excel
from starlette.middleware.sessions import SessionMiddleware

from models import VehicleRecord

# ---------- Config ----------
DATABASE_URL = os.getenv("DATABASE_URL")
engine = create_engine(DATABASE_URL, echo=False)


# Read OpenRouter key from environment (safer)
OPENROUTER_API_KEY = os.getenv("OPENROUTER_API_KEY")
APP_USERNAME = os.getenv("APP_USERNAME")
APP_PASSWORD = os.getenv("APP_PASSWORD")

# Session lifetime (seconds) before forcing re-login
SESSION_TIMEOUT_SECONDS = 5 * 60  # e.g. 5 minutes


# ---------- App & Templates ----------
app = FastAPI(title="Tractor Insurance Records")

firebase_key = os.getenv("FIREBASE_PRIVATE_KEY").replace("\\n", "\n")

cred = credentials.Certificate({
    "type": "service_account",
    "project_id": "tractorcare-8586f",
    "private_key": firebase_key,
    "client_email": os.getenv("FIREBASE_CLIENT_EMAIL"),
    "token_uri": "https://oauth2.googleapis.com/token",
})

firebase_admin.initialize_app(cred)


# Sessions for login
app.add_middleware(
    SessionMiddleware,
    secret_key=os.getenv("SESSION_SECRET_KEY", "fallback-secret-key"),
  # CHANGE THIS AS WELL
    session_cookie="tractorcare_session",
)

app.mount("/static", StaticFiles(directory="static"), name="static")
templates = Jinja2Templates(directory="templates")
scheduler = BackgroundScheduler()


# ---------- DB helpers ----------
def create_db_and_tables():
    SQLModel.metadata.create_all(engine)


def get_session():
    with Session(engine) as session:
        yield session


# ---------- Reminder job ----------
def check_expiring_policies():
    today = date.today()
    with Session(engine) as session:
        for r in session.exec(select(VehicleRecord)).all():
            days = (r.validity_date - today).days
            if days == 5 and r.device_token:
                send_push(
                    r.device_token,
                    "🚨 Policy Expiring",
                    f"{r.vehicle_model} expires in 5 days!"
                )



# ---------- OCR helpers ----------
def extract_text_from_bytes(raw_bytes: bytes, filename: str) -> str:
    filename = filename.lower()

    # Images: png, jpg, jpeg, webp
    if filename.endswith((".png", ".jpg", ".jpeg", ".webp")):
        img = Image.open(io.BytesIO(raw_bytes)).convert("RGB")
        return pytesseract.image_to_string(img)

    # Excel: xlsx, xlsm
    if filename.endswith((".xlsx", ".xlsm")):
        wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), data_only=True)
        text = ""
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for row in ws.iter_rows(values_only=True):
                text += " | ".join([str(c) for c in row if c is not None]) + "\n"
        return text

    # Fallback: try plain text
    try:
        return raw_bytes.decode("utf-8", errors="ignore")
    except Exception:
        return ""


# ---------- Improved Excel parser helpers ----------
def try_parse_date_string(s: str):
    """Try common date formats, return datetime.date or None."""
    s = s.strip()
    if not s:
        return None
    fmts = (
        "%d.%m.%Y",
        "%d.%m.%y",
        "%d/%m/%Y",
        "%d/%m/%y",
        "%Y-%m-%d",
        "%d-%m-%Y",
        "%m/%d/%Y",
    )
    for fmt in fmts:
        try:
            return datetime.datetime.strptime(s, fmt).date()
        except Exception:
            pass
    # ISO fallback
    try:
        return datetime.date.fromisoformat(s)
    except Exception:
        pass
    # regex dd/mm/yy or dd.mm.yy fallback
    m = re.search(r"(\d{1,2})[./-](\d{1,2})[./-](\d{2,4})", s)
    if m:
        day, mon, yr = m.group(1), m.group(2), m.group(3)
        if len(yr) == 2:
            yr = "20" + yr
        try:
            return datetime.date(int(yr), int(mon), int(day))
        except Exception:
            pass
    return None


def extract_end_date_from_cell(cell):
    """
    Given a cell (which may be datetime, date, number, or string like
    'DD.MM.YYYY TO DD.MM.YYYY'), return an ISO date string for the *end*
    date, or empty string if not parseable.
    """
    if cell is None:
        return ""
    # Excel datetime / date objects
    if isinstance(cell, datetime.datetime):
        return cell.date().isoformat()
    if isinstance(cell, datetime.date):
        return cell.isoformat()
    # If numeric (Excel serial)
    if isinstance(cell, (int, float)):
        try:
            dt = openpyxl_from_excel(cell)
            if isinstance(dt, datetime.datetime):
                return dt.date().isoformat()
            if isinstance(dt, datetime.date):
                return dt.isoformat()
        except Exception:
            pass
    # string handling
    s = str(cell).strip()
    # "to" range
    if re.search(r"\bto\b", s, flags=re.IGNORECASE):
        parts = re.split(r"\bto\b", s, flags=re.IGNORECASE)
        if len(parts) >= 2:
            cand = parts[1].strip()
            d = try_parse_date_string(cand)
            if d:
                return d.isoformat()
    # dash range: "DD.MM.YYYY - DD.MM.YYYY"
    if "-" in s and len(s.split("-")) >= 2:
        parts = [p.strip() for p in s.split("-")]
        cand = parts[-1]
        d = try_parse_date_string(cand)
        if d:
            return d.isoformat()
    # single date
    d = try_parse_date_string(s)
    if d:
        return d.isoformat()
    return ""


def parse_excel_rows(raw_bytes: bytes, filename: str):
    """
    Parse Excel workbook and return list of rows as dicts.
    This function is robust to different header names and date formats.
    """
    wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    # Normalize headers (lowercased)
    headers = [str(h).strip().lower() if h is not None else "" for h in rows[0]]

    # Build a mapping from column index -> standard key
    key_map = {}
    for i, h in enumerate(headers):
        if "model" in h:
            key_map[i] = "vehicle_model"
        elif any(x in h for x in ("r.c", "rc", "rc no", "r.c. no", "r.c. no.")) or "register" in h:
            key_map[i] = "register_number"
        elif "name" in h and "ins" not in h:
            key_map[i] = "owner_name"
        elif "address" in h:
            key_map[i] = "address"
        elif ("valid" in h and "date" in h) or "validity" in h:
            key_map[i] = "validity_date"
        elif "policy" in h and "num" in h:
            key_map[i] = "policy_number"
        elif "insurance" in h:
            key_map[i] = "insurance_company"
        else:
            # unknown header — ignore it
            pass

    results = []
    # iterate rows after header row
    for r in rows[1:]:
        entry = {
            "register_number": "",
            "vehicle_model": "",
            "owner_name": "",
            "policy_number": "",
            "validity_date": "",
            "address": "",
            "insurance_company": "",
        }
        for i, cell in enumerate(r):
            if i not in key_map:
                continue
            key = key_map[i]
            if key == "validity_date":
                entry[key] = extract_end_date_from_cell(cell)
            else:
                entry[key] = "" if cell is None else str(cell).strip()
        results.append(entry)
    return results


# ---------- Heuristic insurance guess ----------
def heuristic_insurance_guess(policy_number: str) -> str:
    if not policy_number:
        return ""
    p = policy_number.upper()
    if p.startswith("POL"):
        return "General Insurance Co."
    if p.startswith("IC"):
        return "ICICI Lombard"
    if "BAJAJ" in p or p.startswith("BAJ"):
        return "Bajaj Allianz"
    if "HDFC" in p:
        return "HDFC Ergo"
    if "RAJ" in p:
        return "Reliance General"
    # fallback
    return ""


# ---------- OpenRouter AI helper for single-row insurance guess ----------
def ai_guess_insurance_for_row(row_text: str) -> str:
    """
    Call OpenRouter to return a guessed insurance company name for the row.
    Returns company name string or empty string on failure.
    Only used when field is missing.
    """
    if not OPENROUTER_API_KEY:
        print("AI disabled: OPENROUTER_API_KEY is not set.")
        return ""
    prompt = f"""
From this single record text, guess the insurance company name if possible.

Return only the company name (no JSON, no extra text). If unknown, return an empty string.

Text:
{row_text}
"""
    url = "https://openrouter.ai/api/v1/chat/completions"
    headers = {
        "Authorization": f"Bearer {OPENROUTER_API_KEY}",
        "Content-Type": "application/json",
    }
    payload = {
        "model": "gpt-4o-mini",
        "messages": [
            {
                "role": "system",
                "content": "You are an assistant that returns a single insurance company name based on policy text.",
            },
            {"role": "user", "content": prompt},
        ],
        "temperature": 0.0,
        "max_tokens": 60,
    }
    try:
        resp = requests.post(url, headers=headers, json=payload, timeout=20)
    except Exception as e:
        print("AI guess failed (network):", e)
        return ""

    if resp.status_code == 401:
        print("AI guess failed: 401 Unauthorized – check OPENROUTER_API_KEY.")
        return ""
    if resp.status_code >= 400:
        print(f"AI guess failed: {resp.status_code} {resp.text[:200]}")
        return ""

    try:
        data = resp.json()
        assistant_text = data.get("choices", [{}])[0].get("message", {}).get("content", "")
        answer = assistant_text.strip().splitlines()[0].strip()
        if len(answer) > 200:
            return ""
        return answer
    except Exception as e:
        print("AI guess parse failed:", e)
        return ""


# ---------- Auth helper (dependency) ----------
def require_login(request: Request):
    """
    Simple dependency to guard routes.
    If not logged in or session expired -> redirect to /login.
    """
    if "session" not in request.scope:
        # SessionMiddleware not working
        raise HTTPException(
            status_code=status.HTTP_303_SEE_OTHER,
            headers={"Location": "/login"},
        )

    sess = request.session
    if not sess.get("logged_in"):
        raise HTTPException(
            status_code=status.HTTP_303_SEE_OTHER,
            headers={"Location": "/login"},
        )

    now = time.time()
    last_seen = sess.get("last_seen", 0)

    # Timeout handling
    if SESSION_TIMEOUT_SECONDS and last_seen and (now - last_seen) > SESSION_TIMEOUT_SECONDS:
        sess.clear()
        raise HTTPException(
            status_code=status.HTTP_303_SEE_OTHER,
            headers={"Location": "/login"},
        )

    # Update activity timestamp
    sess["last_seen"] = now
    return True

def send_push(token, title, body):
    msg = messaging.Message(
        token=token,
        notification=messaging.Notification(
            title=title,
            body=body,
        )
    )
    messaging.send(msg)



# ---------- Login / Logout ----------
@app.get("/login", response_class=HTMLResponse)
def login_form(request: Request):
    # Always show login page; user must log in again each time session expired
    return templates.TemplateResponse("login.html", {"request": request, "error": None})


@app.post("/login", response_class=HTMLResponse)
async def login_submit(
    request: Request,
    username: str = Form(...),
    password: str = Form(...),
):
    if username == APP_USERNAME and password == APP_PASSWORD:
        # fresh session
        request.session.clear()
        request.session["logged_in"] = True
        request.session["user"] = username
        request.session["last_seen"] = time.time()
        return RedirectResponse("/", status_code=status.HTTP_303_SEE_OTHER)

    return templates.TemplateResponse(
        "login.html",
        {"request": request, "error": "Invalid username or password."},
    )


@app.get("/logout")
def logout(request: Request):
    if "session" in request.scope:
        request.session.clear()
    return RedirectResponse("/login", status_code=status.HTTP_303_SEE_OTHER)


# ---------- Startup / Shutdown ----------
@app.on_event("startup")
def on_startup():
    # Only allow ONE worker to perform migrations
    if os.environ.get("RUN_MAIN") == "true" or os.environ.get("RENDER") == "true":
        create_db_and_tables()

    try:
        scheduler.add_job(check_expiring_policies, "interval", hours=1)
        scheduler.start()
        print("Reminder scheduler started")

    except Exception:
        pass



@app.on_event("shutdown")
def on_shutdown():
    try:
        scheduler.shutdown()
    except Exception:
        pass


# ---------- Routes: basic CRUD (all protected with require_login) ----------
@app.get("/", response_class=HTMLResponse)
def dashboard(
    request: Request,
    session: Session = Depends(get_session),
    _user: bool = Depends(require_login),
):
    records = session.exec(select(VehicleRecord)).all()
    today = date.today()
    enriched = []
    for r in records:
        days_left = (r.validity_date - today).days
        if days_left < 0:
            status = "Expired"
        elif days_left <= 7:
            status = "Urgent"
        elif days_left <= 30:
            status = "Soon"
        else:
            status = "OK"
        enriched.append({"record": r, "days_left": days_left, "status": status})
    return templates.TemplateResponse(
        "index.html",
        {"request": request, "records": enriched, "today": today},
    )


@app.get("/records/new", response_class=HTMLResponse)
def new_record_form(
    request: Request,
    _user: bool = Depends(require_login),
):
    return templates.TemplateResponse("new_record.html", {"request": request})

@app.post("/register-device")
async def register_device(request: Request, session: Session = Depends(get_session)):
    data = await request.json()
    token = data.get("token","").strip()

    if not token:
        print("❌ EMPTY TOKEN RECEIVED")
        return {"ok": False}

    session.add(Device(token=token))
    session.commit()
    print("✅ TOKEN SAVED:", token[:30])
    return {"ok": True}


@app.post("/records/new")
def create_record(
    register_number: str = Form(...),
    vehicle_model: str = Form(...),
    owner_name: str = Form(...),
    address: str = Form(...),
    policy_number: str = Form(...),
    validity_date: str = Form(...),
    insurance_company: str = Form(...),
    session: Session = Depends(get_session),
    _user: bool = Depends(require_login),
    device_token: str = Form(""),

):
    record = VehicleRecord(
        register_number=register_number,
        vehicle_model=vehicle_model,
        owner_name=owner_name,
        address=address,
        policy_number=policy_number,
        validity_date=date.fromisoformat(validity_date),
        insurance_company=insurance_company,
        device_token=device_token,
    )
    session.add(record)
    session.commit()
    return RedirectResponse("/", status_code=status.HTTP_303_SEE_OTHER)


@app.get("/records/{record_id}", response_class=HTMLResponse)
def record_detail(
    record_id: int,
    request: Request,
    session: Session = Depends(get_session),
    _user: bool = Depends(require_login),
):
    record = session.get(VehicleRecord, record_id)
    if not record:
        return RedirectResponse("/", status_code=status.HTTP_303_SEE_OTHER)
    return templates.TemplateResponse("record_detail.html", {"request": request, "record": record})


@app.get("/records/{record_id}/edit", response_class=HTMLResponse)
def edit_record_form(
    record_id: int,
    request: Request,
    session: Session = Depends(get_session),
    _user: bool = Depends(require_login),
):
    record = session.get(VehicleRecord, record_id)
    if not record:
        raise HTTPException(status_code=404, detail="Record not found")
    return templates.TemplateResponse("edit_record.html", {"request": request, "record": record, "back": "/"})


@app.post("/records/{record_id}/edit")
def edit_record_submit(
    record_id: int,
    register_number: str = Form(...),
    vehicle_model: str = Form(...),
    owner_name: str = Form(...),
    address: str = Form(...),
    policy_number: str = Form(...),
    validity_date: str = Form(...),
    insurance_company: str = Form(...),
    session: Session = Depends(get_session),
    _user: bool = Depends(require_login),
    device_token: str = Form(""),

):
    record = session.get(VehicleRecord, record_id)
    if not record:
        raise HTTPException(status_code=404, detail="Record not found")
    record.register_number = register_number
    record.vehicle_model = vehicle_model
    record.owner_name = owner_name
    record.address = address
    record.policy_number = policy_number
    record.validity_date = date.fromisoformat(validity_date)
    record.insurance_company = insurance_company
    record.device_token = device_token
    session.add(record)
    session.commit()
    return RedirectResponse(f"/records/{record_id}", status_code=status.HTTP_303_SEE_OTHER)


@app.post("/records/{record_id}/delete")
def delete_record(
    record_id: int,
    session: Session = Depends(get_session),
    _user: bool = Depends(require_login),
):
    record = session.get(VehicleRecord, record_id)
    if not record:
        raise HTTPException(status_code=404, detail="Record not found")
    session.delete(record)
    session.commit()
    return RedirectResponse("/", status_code=status.HTTP_303_SEE_OTHER)


# serve service-worker if present
@app.get("/service-worker.js", include_in_schema=False)
def service_worker():
    sw_file = Path(__file__).with_name("service-worker.js")
    return FileResponse(sw_file)


# ---------- Excel / OCR multi-upload endpoints ----------
@app.get("/ocr-upload", response_class=HTMLResponse)
def upload_page(
    request: Request,
    _user: bool = Depends(require_login),
):
    return templates.TemplateResponse("ocr_upload.html", {"request": request})


@app.post("/ocr-upload")
async def ocr_upload(
    request: Request,
    file: UploadFile = File(...),
    _user: bool = Depends(require_login),
):
    """
    Detect file type. If excel -> parse rows and show multi-result preview.
    If image/pdf -> run single-record OCR/AI flow.
    """
    raw = await file.read()
    filename = file.filename.lower()

    # Excel path: parse rows and show multi preview
    if filename.endswith((".xlsx", ".xlsm")):
        entries = parse_excel_rows(raw, filename)
        # apply heuristic guesses for insurance_company where missing
        for e in entries:
            if not e.get("insurance_company"):
                guess = heuristic_insurance_guess(e.get("policy_number", ""))
                e["_insurance_heuristic"] = guess
                e["insurance_company"] = guess or ""
            else:
                e["_insurance_heuristic"] = e["insurance_company"]
        return templates.TemplateResponse("ocr_multi_result.html", {"request": request, "entries": entries})

    # non-excel: do single-record OCR -> AI
    extracted_text = extract_text_from_bytes(raw, filename)
    try:
        fields = ai_extract_fields_openrouter(extracted_text)
    except Exception as e:
        print("AI failed for full doc:", e)
        fields = {
            "register_number": "",
            "vehicle_model": "",
            "owner_name": "",
            "address": "",
            "policy_number": "",
            "validity_date": "",
            "insurance_company": "",
        }
    return templates.TemplateResponse("ocr_result.html", {"request": request, "fields": fields})


# ---------- AI-fill endpoint for excel rows ----------
@app.post("/excel-ai-fill")
async def excel_ai_fill(
    request: Request,
    _user: bool = Depends(require_login),
):
    """
    Accept JSON body: { rows: [ { register_number, vehicle_model, ... } ] }
    Returns same list with 'ai_insurance' filled where possible (only missing ones).
    """
    body = await request.json()
    rows = body.get("rows", [])
    results = []
    for r in rows:
        if r.get("insurance_company"):
            r["ai_insurance"] = r.get("insurance_company")
            results.append(r)
            continue
        row_text = " | ".join([f"{k}:{v}" for k, v in r.items() if v])
        ai_guess = ai_guess_insurance_for_row(row_text)
        r["ai_insurance"] = ai_guess
        results.append(r)
    return JSONResponse({"rows": results})


# ---------- Save individual row ----------
@app.post("/records/add_row")
def add_single_row(
    register_number: str = Form(...),
    vehicle_model: str = Form(...),
    owner_name: str = Form(...),
    address: str = Form(...),
    policy_number: str = Form(...),
    validity_date: str = Form(...),
    insurance_company: str = Form(...),
    session: Session = Depends(get_session),
    _user: bool = Depends(require_login),
):
    record = VehicleRecord(
        register_number=register_number,
        vehicle_model=vehicle_model,
        owner_name=owner_name,
        address=address,
        policy_number=policy_number,
        validity_date=date.fromisoformat(validity_date),
        insurance_company=insurance_company,
    )
    session.add(record)
    session.commit()
    return RedirectResponse("/", status_code=status.HTTP_303_SEE_OTHER)


# ---------- Save all rows endpoint ----------
@app.post("/records/add_rows_bulk")
async def add_rows_bulk(
    request: Request,
    session: Session = Depends(get_session),
    _user: bool = Depends(require_login),
):
    body = await request.form()
    rows_json = body.get("rows_json")
    if not rows_json:
        raise HTTPException(status_code=400, detail="Missing rows_json")
    rows = json.loads(rows_json)
    for r in rows:
        try:
            vd = date.fromisoformat(r.get("validity_date"))
        except Exception:
            continue
        rec = VehicleRecord(
            register_number=r.get("register_number", ""),
            vehicle_model=r.get("vehicle_model", ""),
            owner_name=r.get("owner_name", ""),
            address=r.get("address", ""),
            policy_number=r.get("policy_number", ""),
            validity_date=vd,
            insurance_company=r.get("insurance_company", ""),
        )
        session.add(rec)
    session.commit()
    return RedirectResponse("/", status_code=status.HTTP_303_SEE_OTHER)

@app.get("/firebase-messaging-sw.js", include_in_schema=False)
def fcm_worker():
    return FileResponse("firebase-messaging-sw.js")

@app.get("/token", response_class=HTMLResponse)
def get_my_token(request: Request):
    return HTMLResponse("""
<!DOCTYPE html>
<html>
<head>
<title>TractorCare Token</title>
<script src="https://www.gstatic.com/firebasejs/9.23.0/firebase-app-compat.js"></script>
<script src="https://www.gstatic.com/firebasejs/9.23.0/firebase-messaging-compat.js"></script>
</head>
<body style="background:black;color:white;font-family:monospace;padding:20px">

<h2>TractorCare Device Token</h2>
<p id="status">Waiting for permission…</p>
<pre id="token" style="white-space:break-spaces;"></pre>

<script>
const firebaseConfig = {
  apiKey: "AIzaSyBYAuUVkA9oOHon2i93jZPYOCxxlJ8jvyM",
  authDomain: "tractorcare-8586f.firebaseapp.com",
  projectId: "tractorcare-8586f",
  messagingSenderId: "697692026188",
  appId: "1:697692026188:web:3dafe18a7b6aa8a1794be9"
};

firebase.initializeApp(firebaseConfig);
const messaging = firebase.messaging();

Notification.requestPermission().then(async perm=>{
  if(perm==="granted"){
    const tok = await messaging.getToken({ vapidKey:"BGd2pZ6QoNSygh2yPG2d_mUMEEiT7EEXlVjNlAfham7q5uIYVH3Zoudqoc0aMxTigWRqXlorXB9Fex79rbKDipc"});
    document.getElementById("token").innerText = tok;
    document.getElementById("status").innerText = "TOKEN GENERATED:";
  } else {
    document.getElementById("status").innerText = "Permission denied.";
  }
});
</script>
</body>
</html>
""")

@app.get("/test-push")
def test_push():
    from sqlmodel import Session
    with Session(engine) as session:
        r = session.exec(select(VehicleRecord)).first()
        if not r or not r.device_token:
            return {"error": "No record with device token found"}

        send_push(
            r.device_token,
            "🔥 TractorCare Test",
            "This is your first real Android notification!"
        )
        return {"ok": True}


# ---------- single-record AI extraction ----------
def ai_extract_fields_openrouter(text: str) -> dict:
    """
    Return JSON with the full set of fields when called for single documents.
    """
    if not OPENROUTER_API_KEY:
        raise RuntimeError("OPENROUTER_API_KEY not set")
    prompt = f"""
Extract the following fields from this vehicle insurance or service document text:

- register_number
- vehicle_model
- owner_name
- address
- policy_number
- validity_date
- insurance_company

Return ONLY valid JSON like this:

{{
  "register_number": "",
  "vehicle_model": "",
  "owner_name": "",
  "address": "",
  "policy_number": "",
  "validity_date": "",
  "insurance_company": ""
}}

If you can't find a field, keep it empty.

TEXT:
{text}
"""
    url = "https://openrouter.ai/api/v1/chat/completions"
    headers = {"Authorization": f"Bearer {OPENROUTER_API_KEY}", "Content-Type": "application/json"}
    payload = {
        "model": "gpt-4o-mini",
        "messages": [
            {"role": "system", "content": "You are a JSON extraction assistant. Respond with JSON only."},
            {"role": "user", "content": prompt},
        ],
        "temperature": 0,
        "max_tokens": 700,
    }
    try:
        resp = requests.post(url, headers=headers, json=payload, timeout=30)
    except Exception as e:
        print("OpenRouter request failed (network):", e)
        return {
            k: ""
            for k in [
                "register_number",
                "vehicle_model",
                "owner_name",
                "address",
                "policy_number",
                "validity_date",
                "insurance_company",
            ]
        }

    if resp.status_code == 401:
        print("OpenRouter request failed: 401 Unauthorized – check OPENROUTER_API_KEY.")
        return {
            k: ""
            for k in [
                "register_number",
                "vehicle_model",
                "owner_name",
                "address",
                "policy_number",
                "validity_date",
                "insurance_company",
            ]
        }
    if resp.status_code >= 400:
        print(f"OpenRouter request failed: {resp.status_code} {resp.text[:200]}")
        return {
            k: ""
            for k in [
                "register_number",
                "vehicle_model",
                "owner_name",
                "address",
                "policy_number",
                "validity_date",
                "insurance_company",
            ]
        }

    try:
        data = resp.json()
        assistant_text = data.get("choices", [{}])[0].get("message", {}).get("content", "")
        match = re.search(r"(\{[\s\S]*?\})", assistant_text)
        json_text = match.group(1) if match else assistant_text
        parsed = json.loads(json_text)
    except Exception as e:
        print("OpenRouter JSON parse failed:", e)
        parsed = {}

    keys = [
        "register_number",
        "vehicle_model",
        "owner_name",
        "address",
        "policy_number",
        "validity_date",
        "insurance_company",
    ]
    return {k: str(parsed.get(k, "") if parsed.get(k, "") is not None else "") for k in keys}
